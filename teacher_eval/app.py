import os, sys, io, random
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, Admin, Evaluation, ClassConfig, QuestionConfig

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'gaozhou-erzhong-eval-secret-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///evaluation.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'admin_login'


@login_manager.user_loader
def load_user(user_id):
    return Admin.query.get(int(user_id))


DEFAULT_QUESTIONS = [
    "该教师遵守教师职业道德规范，如不体罚、不歧视学生等。",
    "该老师的工作责任心强，关心爱护学生。",
    "该老师从不组织学生征订除\"一教一辅\"外的教辅资料。",
    "该老师能够按时上下课和值班。",
    "该老师使用普通话授课。",
    "该老师经常在学科教学中融入思政教育，帮助学生树立正确的世界观、人生观和价值观。",
    "该老师的课堂组织管理严格、有序，同学们基本都能认真听课。",
    "该老师备课充分，教学内容清晰、有条理。",
    "该老师上课生动形象，重视基础和思维拓展，重难点突出。",
    "该老师经常在课堂上对同学们进行学习方法和考试方法指导。",
    "老师在课堂上注重师生互动，积极提问，善于调动学生积极性。",
    "该老师的上课节奏合适，大部分同学可以适应。",
    "该老师讲课的声音悦耳，音量层次分明。",
    "该老师在上课时认真板书，而且板书清楚、美观、有条理（体育老师选A）。",
    "该老师布置学生利用白天课堂时间自习时，学习任务明确，并且能亲自坐班管理。",
    "该老师能合理布置课后作业，批改认真，讲评及时，讲评效果好（非九大科老师选A）。",
    "该老师能按时布置日测和周测，批改认真，讲评及时，讲评效果好（非九大科老师选A）。",
    "该老师早晚读、自修或周末值班工作认真、勤管纪律、耐心辅导学生。",
    "该老师经常对学生进行个别谈心、个别辅导和思想教育。",
    "你对该教师的总体评价非常满意。",
]

SUBJECTS = ['语文','数学','英语','物理','化学','生物','政治','历史','地理','体育','信息技术','音乐','美术','心理']


def get_questions():
    qs = QuestionConfig.query.order_by(QuestionConfig.question_number).all()
    if len(qs) >= 20:
        return [q.content for q in qs[:20]]
    return DEFAULT_QUESTIONS


def get_question_configs():
    qs = QuestionConfig.query.order_by(QuestionConfig.question_number).all()
    if len(qs) >= 20:
        return qs[:20]
    return None


def get_max_score():
    qs = get_question_configs()
    if qs:
        return sum(q.max_score for q in qs)
    return 100


def avg_score_expr():
    """Correct SQL expression: average of all question scores per evaluation row."""
    qs = get_question_configs()
    qn = len(qs) if qs and len(qs) >= 20 else 20
    total = sum(qs[i].max_score for i in range(qn)) if qs else 100
    case_sum = sum(
        db.case((getattr(Evaluation, f'q{i+1}') == 'A', qs[i].max_score if qs else 5),
                (getattr(Evaluation, f'q{i+1}') == 'B', round((qs[i].max_score if qs else 5) * 0.8)),
                (getattr(Evaluation, f'q{i+1}') == 'C', round((qs[i].max_score if qs else 5) * 0.6)),
                (getattr(Evaluation, f'q{i+1}') == 'D', round((qs[i].max_score if qs else 5) * 0.4)),
                (getattr(Evaluation, f'q{i+1}') == 'E', round((qs[i].max_score if qs else 5) * 0.2)),
                else_=0)
        for i in range(qn)
    )
    return case_sum / qn


def init_db():
    with app.app_context():
        # migrate: drop old single-column unique index on device_uuid so one device can eval multiple subjects
        try:
            inspector = db.inspect(db.engine)
            for ix in inspector.get_indexes('evaluation'):
                if ix.get('unique') and ix['column_names'] == ['device_uuid']:
                    db.session.execute(db.text(f'DROP INDEX "{ix["name"]}"'))
                    db.session.commit()
        except Exception:
            db.session.rollback()
        db.create_all()
        if not Admin.query.filter_by(username='admin').first():
            a = Admin(username='admin'); a.set_password('admin123')
            db.session.add(a)
        for g in ['高一','高二','高三']:
            if not ClassConfig.query.filter_by(grade=g).first():
                db.session.add(ClassConfig(grade=g, class_count=33 if g != '高三' else 0))
        if QuestionConfig.query.count() == 0:
            for i, txt in enumerate(DEFAULT_QUESTIONS, 1):
                db.session.add(QuestionConfig(question_number=i, content=txt))
        db.session.commit()
        print('DB initialized: admin/admin123 | class configs | questions seeded')


# ─── API ───────────────────────────────────────────────
@app.route('/api/grades')
def api_grades():
    configs = ClassConfig.query.order_by(ClassConfig.id).all()
    return jsonify([{'value': c.grade, 'label': c.grade} for c in configs if c.class_count > 0])


@app.route('/api/classes/<grade>')
def api_classes(grade):
    count = ClassConfig.get_count(grade)
    return jsonify([{'value': str(i), 'label': f'{i}班'} for i in range(1, count + 1)])


@app.route('/api/my-evaluation')
def api_my_evaluation():
    uuid = request.args.get('uuid', '')
    if not uuid:
        return jsonify([])
    evals = Evaluation.query.filter_by(device_uuid=uuid).order_by(Evaluation.created_at.desc()).all()
    return jsonify([{
        'grade': e.grade, 'class_name': e.class_name + '班',
        'teacher_subject': e.teacher_subject,
        'submitted': e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else ''
    } for e in evals])


@app.route('/api/questions')
def api_questions():
    return jsonify(get_questions_with_id())


# ─── Survey ────────────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
def survey():
    questions = get_questions()
    if request.method == 'POST':
        grade = request.form.get('grade', '')
        class_name = request.form.get('className', '')
        teacher_subject = request.form.get('teacherSubject', '')
        device_uuid = request.form.get('device_uuid', '')

        if not all([grade, class_name, teacher_subject]):
            flash('请完整填写基本信息', 'error')
            return render_template('survey_original.html', questions=questions, q_count=len(questions))

        for i in range(1, len(questions) + 1):
            if not request.form.get(f'q{i}', ''):
                flash(f'请完成第{i}题的评价', 'error')
                return render_template('survey_original.html', questions=questions, q_count=len(questions))

        suggestions = request.form.get('suggestions', '')
        ip_address = request.remote_addr or ''

        # upsert by (device_uuid, teacher_subject) — pure app-level, no DB constraint needed
        existing = None
        if device_uuid and teacher_subject:
            existing = Evaluation.query.filter_by(device_uuid=device_uuid, teacher_subject=teacher_subject).first()

        if existing:
            existing.grade = grade; existing.class_name = class_name
            existing.teacher_subject = teacher_subject
            for i in range(1, len(questions) + 1):
                setattr(existing, f'q{i}', request.form.get(f'q{i}', ''))
            existing.suggestions = suggestions; existing.ip_address = ip_address
            existing.updated_at = datetime.now()
            db.session.commit()
            return redirect(url_for('success', id=existing.id))

        eval_data = {'grade': grade, 'class_name': class_name,
                     'teacher_subject': teacher_subject, 'device_uuid': device_uuid or None,
                     'suggestions': suggestions, 'ip_address': ip_address}
        for i in range(1, len(questions) + 1):
            eval_data[f'q{i}'] = request.form.get(f'q{i}', '')
        evaluation = Evaluation(**eval_data)
        db.session.add(evaluation)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
            # race condition: another submission came in between check and insert
            existing = Evaluation.query.filter_by(device_uuid=device_uuid, teacher_subject=teacher_subject).first()
            if existing:
                for k, v in eval_data.items():
                    setattr(existing, k, v)
                existing.updated_at = datetime.now()
                db.session.commit()
                return redirect(url_for('success', id=existing.id))
            raise

        return redirect(url_for('success', id=evaluation.id))

    return render_template('survey_original.html', questions=questions, q_count=len(questions))


@app.route('/success')
def success():
    eval_id = request.args.get('id', 0, type=int)
    evaluation = Evaluation.query.get_or_404(eval_id) if eval_id else None
    return render_template('success.html', evaluation=evaluation)


# ─── Admin login ───────────────────────────────────────
@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if current_user.is_authenticated:
        return redirect(url_for('admin_dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        admin = Admin.query.filter_by(username=username).first()
        if admin and admin.check_password(password):
            login_user(admin)
            return redirect(url_for('admin_dashboard'))
        flash('用户名或密码错误', 'error')
    return render_template('admin/login.html')


@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    return redirect(url_for('admin_login'))


# ─── Admin: accounts ───────────────────────────────────
@app.route('/admin/admins', methods=['GET', 'POST'])
@login_required
def admin_admins():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or len(password) < 4:
            flash('用户名不能为空，密码至少4位', 'error')
        elif Admin.query.filter_by(username=username).first():
            flash('用户名已存在', 'error')
        else:
            a = Admin(username=username); a.set_password(password)
            db.session.add(a); db.session.commit()
            flash(f'管理员「{username}」已添加', 'success')
        return redirect(url_for('admin_admins'))
    admins = Admin.query.order_by(Admin.id).all()
    return render_template('admin/admins.html', admins=admins)


@app.route('/admin/admins/delete/<int:id>', methods=['POST'])
@login_required
def admin_delete_admin(id):
    a = Admin.query.get_or_404(id)
    if a.id == current_user.id:
        flash('不能删除自己', 'error')
    elif a.username == 'admin':
        flash('不能删除初始管理员', 'error')
    else:
        db.session.delete(a); db.session.commit()
        flash(f'已删除管理员「{a.username}」', 'success')
    return redirect(url_for('admin_admins'))


# ─── Admin: questions ──────────────────────────────────
@app.route('/admin/questions', methods=['GET', 'POST'])
@login_required
def admin_questions():
    if request.method == 'POST':
        qid = request.form.get('qid', 0, type=int)
        content = request.form.get('content', '').strip()
        if qid and content:
            q = QuestionConfig.query.get(qid)
            if q:
                q.content = content
                db.session.commit()
                flash(f'第{q.question_number}题已更新', 'success')
        return redirect(url_for('admin_questions'))
    questions = QuestionConfig.query.order_by(QuestionConfig.question_number).all()
    return render_template('admin/questions.html', questions=questions)


# ─── Admin dashboard ──────────────────────────────────
@app.route('/admin')
@login_required
def admin_dashboard():
    total = Evaluation.query.count()
    unique_devices = db.session.query(db.func.count(db.distinct(Evaluation.device_uuid))).scalar() or 0
    grades = db.session.query(Evaluation.grade, db.func.count(Evaluation.id)).group_by(Evaluation.grade).all()
    subjects = db.session.query(Evaluation.teacher_subject, db.func.count(Evaluation.id)).group_by(Evaluation.teacher_subject).all()
    recent = Evaluation.query.order_by(Evaluation.created_at.desc()).limit(10).all()
    q_count = len(get_questions())

    all_scores = [e.average_score for e in Evaluation.query.all()]
    overall_avg = round(sum(all_scores) / len(all_scores), 2) if all_scores else 0
    dist = {'高 (80-100分)': 0, '中 (60-80分)': 0, '低 (0-60分)': 0}
    for s in all_scores:
        t = s * q_count
        if t >= 80: dist['高 (80-100分)'] += 1
        elif t >= 60: dist['中 (60-80分)'] += 1
        else: dist['低 (0-60分)'] += 1

    today = datetime.now().date()
    daily_counts = []
    for d in range(6, -1, -1):
        day = today - timedelta(days=d)
        ds = datetime(day.year, day.month, day.day, 0, 0, 0)
        de = datetime(day.year, day.month, day.day, 23, 59, 59)
        cnt = Evaluation.query.filter(Evaluation.created_at >= ds, Evaluation.created_at <= de).count()
        daily_counts.append({'date': day.strftime('%m-%d'), 'count': cnt})

    score_sum = sum(
        db.case((getattr(Evaluation, f'q{i+1}') == 'A', 5),
                (getattr(Evaluation, f'q{i+1}') == 'B', 4),
                (getattr(Evaluation, f'q{i+1}') == 'C', 3),
                (getattr(Evaluation, f'q{i+1}') == 'D', 2),
                (getattr(Evaluation, f'q{i+1}') == 'E', 1), else_=0)
        for i in range(q_count)
    )
    subject_ranking = db.session.query(
        Evaluation.teacher_subject, db.func.count(Evaluation.id).label('cnt'),
        db.func.avg(score_sum / q_count).label('avg')
    ).group_by(Evaluation.teacher_subject).order_by(db.desc('avg')).all()

    avg_scores = {}
    for i in range(1, q_count + 1):
        col = getattr(Evaluation, f'q{i}')
        score_expr = db.case({col == 'A': 5, col == 'B': 4, col == 'C': 3, col == 'D': 2, col == 'E': 1}, else_=0)
        avg = db.session.query(db.func.avg(score_expr)).scalar()
        avg_scores[i] = round(float(avg), 2) if avg else 0

    return render_template('admin/dashboard.html', total=total, unique_devices=unique_devices,
                           grades=grades, subjects=subjects, recent=recent, overall_avg=overall_avg,
                           score_dist=dist, daily_counts=daily_counts, subject_ranking=subject_ranking,
                           avg_scores=avg_scores, questions=get_questions(), q_count=q_count)


# ─── Admin: password ───────────────────────────────────
@app.route('/admin/password', methods=['GET', 'POST'])
@login_required
def admin_password():
    if request.method == 'POST':
        old = request.form.get('old_password', '')
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if not current_user.check_password(old):
            flash('原密码错误', 'error')
        elif len(new) < 4:
            flash('新密码至少4位', 'error')
        elif new != confirm:
            flash('两次输入的新密码不一致', 'error')
        else:
            current_user.set_password(new); db.session.commit()
            flash('密码修改成功', 'success')
            return redirect(url_for('admin_dashboard'))
    return render_template('admin/change_password.html')


# ─── Admin: class configs ──────────────────────────────
@app.route('/admin/classes', methods=['GET', 'POST'])
@login_required
def admin_classes():
    if request.method == 'POST':
        grade = request.form.get('grade', '').strip()
        count = request.form.get('count', 0, type=int)
        if grade and count >= 0:
            cfg = ClassConfig.query.filter_by(grade=grade).first()
            if cfg: cfg.class_count = count
            else: db.session.add(ClassConfig(grade=grade, class_count=count))
            db.session.commit()
            flash(f'{grade}班级数已更新为 {count}', 'success')
        else:
            flash('参数无效', 'error')
        return redirect(url_for('admin_classes'))
    configs = ClassConfig.query.order_by(ClassConfig.id).all()
    return render_template('admin/classes.html', configs=configs)


@app.route('/admin/classes/delete/<grade>', methods=['POST'])
@login_required
def admin_delete_class(grade):
    cfg = ClassConfig.query.filter_by(grade=grade).first()
    if cfg: db.session.delete(cfg); db.session.commit(); flash(f'已删除年级：{grade}', 'success')
    else: flash('该年级不存在', 'error')
    return redirect(url_for('admin_classes'))


# ─── Admin: stats ──────────────────────────────────────
@app.route('/admin/stats')
@login_required
def admin_stats():
    group_by = request.args.get('group', 'grade_class')
    subject_filter = request.args.get('subject', '')
    q_count = len(get_questions())
    score_sum = sum(
        db.case((getattr(Evaluation, f'q{i+1}') == 'A', 5),
                (getattr(Evaluation, f'q{i+1}') == 'B', 4),
                (getattr(Evaluation, f'q{i+1}') == 'C', 3),
                (getattr(Evaluation, f'q{i+1}') == 'D', 2),
                (getattr(Evaluation, f'q{i+1}') == 'E', 1), else_=0)
        for i in range(q_count)
    )
    query = db.session.query(
        Evaluation.grade, Evaluation.class_name, Evaluation.teacher_subject,
        db.func.count(Evaluation.id).label('count'),
        db.func.avg(score_sum / q_count).label('avg_score')
    )
    if subject_filter:
        query = query.filter(Evaluation.teacher_subject == subject_filter)
    groups = {'grade': [Evaluation.grade], 'subject': [Evaluation.teacher_subject],
              'grade_subject': [Evaluation.grade, Evaluation.teacher_subject],
              'grade_class': [Evaluation.grade, Evaluation.class_name, Evaluation.teacher_subject]}
    gcols = groups.get(group_by, groups['grade_class'])
    query = query.group_by(*gcols).order_by(db.desc('avg_score'))
    stats = query.all()
    all_grades = [c.grade for c in ClassConfig.query.order_by(ClassConfig.id).all()]
    return render_template('admin/stats.html', stats=stats, group_by=group_by,
                           subject_filter=subject_filter, subjects=SUBJECTS, all_grades=all_grades)


# ─── Admin: evaluation records ─────────────────────────
@app.route('/admin/results')
@login_required
def admin_results():
    page = request.args.get('page', 1, type=int)
    grade_filter = request.args.get('grade', '')
    class_filter = request.args.get('cls', '')
    subject_filter = request.args.get('subject', '')
    query = Evaluation.query
    if grade_filter: query = query.filter(Evaluation.grade == grade_filter)
    if class_filter: query = query.filter(Evaluation.class_name == class_filter)
    if subject_filter: query = query.filter(Evaluation.teacher_subject == subject_filter)
    query = query.order_by(Evaluation.created_at.desc())
    pagination = query.paginate(page=page, per_page=20, error_out=False)
    all_grades = [c.grade for c in ClassConfig.query.order_by(ClassConfig.id).all()]
    return render_template('admin/results.html', evaluations=pagination.items,
                           pagination=pagination, grade_filter=grade_filter,
                           class_filter=class_filter, subject_filter=subject_filter,
                           all_grades=all_grades)


@app.route('/admin/result/<int:id>')
@login_required
def admin_result_detail(id):
    evaluation = Evaluation.query.get_or_404(id)
    qs = get_questions()
    scores = {i: evaluation.score_of(f'q{i}') for i in range(1, len(qs)+1)}
    return render_template('admin/result_detail.html', evaluation=evaluation,
                           questions=qs, options=dict([('A','完全符合'),('B','基本符合'),
                           ('C','一般'),('D','基本不符合'),('E','完全不符合')]), scores=scores)


@app.route('/admin/result/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def admin_edit_evaluation(id):
    qs = get_questions()
    evaluation = Evaluation.query.get_or_404(id)
    if request.method == 'POST':
        evaluation.grade = request.form.get('grade', evaluation.grade)
        evaluation.class_name = request.form.get('className', evaluation.class_name)
        evaluation.teacher_subject = request.form.get('teacherSubject', evaluation.teacher_subject)
        for i in range(1, len(qs)+1):
            val = request.form.get(f'q{i}', '')
            if val: setattr(evaluation, f'q{i}', val)
        evaluation.suggestions = request.form.get('suggestions', evaluation.suggestions)
        evaluation.updated_at = datetime.now(); db.session.commit()
        flash('评价记录已更新', 'success')
        return redirect(url_for('admin_result_detail', id=evaluation.id))
    all_grades = [c.grade for c in ClassConfig.query.order_by(ClassConfig.id).all()]
    return render_template('admin/edit_evaluation.html', evaluation=evaluation,
                           questions=qs, subjects=SUBJECTS, all_grades=all_grades)


@app.route('/admin/delete/<int:id>', methods=['POST'])
@login_required
def admin_delete(id):
    db.session.delete(Evaluation.query.get_or_404(id)); db.session.commit()
    flash('记录已删除', 'success')
    return redirect(url_for('admin_results'))


@app.route('/admin/clear-all', methods=['POST'])
@login_required
def admin_clear_all():
    count = Evaluation.query.delete()
    db.session.commit()
    flash(f'已清空全部 {count} 条评价记录', 'success')
    return redirect(url_for('admin_dashboard'))


# ─── Export Excel ──────────────────────────────────────
def build_workbook():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = Workbook()
    thin = Side(style='thin', color='d4d4d4')
    hf = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(name='微软雅黑', bold=True, color='ffffff', size=11)
    cfont = Font(name='微软雅黑', size=10)
    align = Alignment(horizontal='center', vertical='center')
    bd = Border(bottom=thin)
    def write_row(ws, row, data, is_header=False):
        for col, v in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = hfont if is_header else cfont
            c.alignment = align
            c.border = bd
            if is_header: c.fill = hf
    def col_letter(n):
        s = ''
        while n > 0: n -= 1; s = chr(65 + n % 26) + s; n //= 26
        return s

    qs = get_questions()
    qs_count = len(qs)

    # Sheet 1: 评价明细
    ws1 = wb.active; ws1.title = '评价明细'
    evals = Evaluation.query.order_by(Evaluation.created_at.desc()).all()
    h1 = ['ID','UUID','年级','班级','学科'] + [f'Q{i}' for i in range(1, qs_count+1)] + \
         [f'Q{i}分' for i in range(1, qs_count+1)] + ['总分','均分','建议','IP','提交时间','更新时间']
    write_row(ws1, 1, h1, True)
    for r, e in enumerate(evals, 2):
        row = [e.id, e.device_uuid or '', e.grade, e.class_name+'班', e.teacher_subject]
        for i in range(1, qs_count+1): row.append(getattr(e, f'q{i}', ''))
        for i in range(1, qs_count+1): row.append(e.score_of(f'q{i}'))
        row += [e.total_score, e.average_score, e.suggestions or '', e.ip_address or '',
                e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '',
                e.updated_at.strftime('%Y-%m-%d %H:%M') if e.updated_at else '']
        write_row(ws1, r, row)
    for i, w in enumerate([6,34,8,8,10]+[8]*qs_count*2+[8,8,30,14,16,16], 1):
        ws1.column_dimensions[col_letter(i)].width = w

    # Sheet 2: 统计汇总
    ws2 = wb.create_sheet('统计汇总')
    h2 = ['年级','班级','学科','评价数','平均分']
    write_row(ws2, 1, h2, True)
    score_sum = sum(
        db.case((getattr(Evaluation, f'q{i+1}') == 'A', 5),
                (getattr(Evaluation, f'q{i+1}') == 'B', 4),
                (getattr(Evaluation, f'q{i+1}') == 'C', 3),
                (getattr(Evaluation, f'q{i+1}') == 'D', 2),
                (getattr(Evaluation, f'q{i+1}') == 'E', 1), else_=0)
        for i in range(qs_count)
    )
    q = db.session.query(Evaluation.grade, Evaluation.class_name, Evaluation.teacher_subject,
        db.func.count(Evaluation.id).label('cnt'),
        db.func.avg(score_sum / qs_count).label('avg')
    ).group_by(Evaluation.grade, Evaluation.class_name, Evaluation.teacher_subject
    ).order_by(db.desc('avg')).all()
    for r, row in enumerate(q, 2):
        write_row(ws2, r, [row.grade, row.class_name+'班', row.teacher_subject, row.cnt, round(float(row.avg or 0), 2)])
    for i, w in enumerate([10,10,12,10,10], 1):
        ws2.column_dimensions[col_letter(i)].width = w

    # Sheet 3: 题目配置
    ws3 = wb.create_sheet('题目配置')
    write_row(ws3, 1, ['题号','题目内容'], True)
    for r, q in enumerate(QuestionConfig.query.order_by(QuestionConfig.question_number).all(), 2):
        write_row(ws3, r, [q.question_number, q.content])
    ws3.column_dimensions['A'].width = 8; ws3.column_dimensions['B'].width = 60

    # Sheet 4: 班级配置
    ws4 = wb.create_sheet('班级配置')
    write_row(ws4, 1, ['年级','班级数','更新时间'], True)
    for r, c in enumerate(ClassConfig.query.order_by(ClassConfig.id).all(), 2):
        write_row(ws4, r, [c.grade, c.class_count,
                    c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else ''])
    for i, w in enumerate([10,10,16], 1): ws4.column_dimensions[col_letter(i)].width = w

    return wb


@app.route('/admin/export')
@login_required
def admin_export():
    wb = build_workbook()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = app.make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename=evaluations_export.xlsx'
    return resp


@app.route('/admin/export/filtered')
@login_required
def admin_export_filtered():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    grade_filter = request.args.get('grade', '')
    class_filter = request.args.get('cls', '')
    subject_filter = request.args.get('subject', '')

    thin = Side(style='thin', color='d4d4d4')
    hf = PatternFill(start_color='1a1a1a', end_color='1a1a1a', fill_type='solid')
    hfont = Font(name='微软雅黑', bold=True, color='ffffff', size=11)
    cfont = Font(name='微软雅黑', size=10)
    align = Alignment(horizontal='center', vertical='center')
    bd = Border(bottom=thin)

    def write_row(ws, row, data, is_header=False):
        for col, v in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = hfont if is_header else cfont; c.alignment = align; c.border = bd
            if is_header: c.fill = hf

    qs = get_questions(); qn = len(qs)
    query = Evaluation.query
    if grade_filter: query = query.filter(Evaluation.grade == grade_filter)
    if class_filter: query = query.filter(Evaluation.class_name == class_filter)
    if subject_filter: query = query.filter(Evaluation.teacher_subject == subject_filter)
    evals = query.order_by(Evaluation.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active; ws.title = '筛选结果'
    h = ['ID','UUID','年级','班级','学科'] + [f'Q{i}' for i in range(1, qn+1)] + \
        [f'Q{i}分' for i in range(1, qn+1)] + ['总分','均分','建议','IP','时间']
    write_row(ws, 1, h, True)
    for r, e in enumerate(evals, 2):
        row = [e.id, e.device_uuid or '', e.grade, e.class_name+'班', e.teacher_subject]
        for i in range(1, qn+1): row.append(getattr(e, f'q{i}', ''))
        for i in range(1, qn+1): row.append(e.score_of(f'q{i}'))
        row += [e.total_score, e.average_score, e.suggestions or '', e.ip_address or '',
                e.created_at.strftime('%Y-%m-%d %H:%M') if e.created_at else '']
        write_row(ws, r, row)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fn_parts = []
    if grade_filter: fn_parts.append(grade_filter)
    if class_filter: fn_parts.append(class_filter + '班')
    if subject_filter: fn_parts.append(subject_filter)
    fn = '_'.join(fn_parts) + '_evaluations.xlsx' if fn_parts else 'evaluations_filtered.xlsx'

    resp = app.make_response(buf.getvalue())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = f'attachment; filename={fn}'
    return resp


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
